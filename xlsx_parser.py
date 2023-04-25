"""Модуль с утилитами для работы с таблицами эксель."""
import os

from openpyxl import load_workbook
from telegram import LabeledPrice


def get_string_shipping_address(shipping_address):
    """Строка адреса доставки из объекта телеграм."""
    if shipping_address:
        shipping_address = ' '.join(
            [
                shipping_address.country_code,
                shipping_address.city,
                shipping_address.street_line1,
                shipping_address.post_code
            ]
        )
        return shipping_address
    return ''


def collect_items(xlsx_file_path, sheet_name):
    """Сбор данных товаров из таблицы эксель и размещение их в памяти."""
    items = dict()
    if os.path.exists(path=xlsx_file_path):
        workbook = load_workbook(xlsx_file_path, read_only=True)
        shop_sheet = workbook[sheet_name]
        for row in shop_sheet.iter_rows(min_row=2):
            if row[0].value:
                items[row[0].value] = {
                    'prices': [LabeledPrice(row[1].value, row[2].value * 100)],
                    'title': row[1].value,
                    'description': row[4].value,
                    'currency': row[3].value
                }
        workbook.close()
        return items


def collect_admins_id(xlsx_file_path, sheet_name):
    """Сбор данных айди админов из таблицы эксель и размещение их в памяти."""
    ids = list()
    if os.path.exists(path=xlsx_file_path):
        workbook = load_workbook(xlsx_file_path, read_only=True)
        admin_sheet = workbook[sheet_name]
        for row in admin_sheet.iter_rows(min_row=2):
            if row[0].value:
                ids.append(row[0].value)
        workbook.close()
        return ids


def add_admin_to_excel(xlsx_file_path, sheet_name, admin_id):
    """Добавление в таблицу эксель данных нового админа."""
    if os.path.exists(path=xlsx_file_path):
        workbook = load_workbook(xlsx_file_path)
        admin_sheet = workbook[sheet_name]
        max_row = admin_sheet.max_row
        coord = 'A' + str(max_row + 1)
        admin_sheet[coord] = admin_id
        workbook.save(xlsx_file_path)
        workbook.close()


def delete_unfilled_rows(sheet):
    """Утилита для удаления рядов таблицы."""
    cur_rows = sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=9)
    idx = list()
    for i in cur_rows:
        if i[0].value:
            continue
        idx.append(i[0].row)
    if idx:
        for i in range(len(idx)):
            sheet.delete_rows(idx[i] - i, 1)
        return sheet
    return sheet


def initiate_transactions_sheet(workbook, sheet_name):
    """Инициализация рабочей таблицы для транзакций."""
    first_row = {
        'A1': 'title',
        'B1': 'price',
        'C1': 'currency',
        'D1': 'name',
        'E1': 'email',
        'F1': 'phone_number',
        'G1': 'shipping_address',
        'H1': 'status',
        'I1': 'datetime'
    }
    workbook.create_sheet(sheet_name)
    transaction_sheet = workbook[sheet_name]
    for cell in first_row:
        transaction_sheet[cell] = first_row.get(cell)
    return transaction_sheet


def add_transaction(xlsx_file_path, sheet_name, **kwargs):
    """Создание записи в таблтице эксель о начале транзакции."""
    if os.path.exists(path=xlsx_file_path):
        workbook = load_workbook(xlsx_file_path)
        if sheet_name not in workbook.sheetnames:
            initiate_transactions_sheet(workbook, sheet_name)
        transaction_sheet = delete_unfilled_rows(workbook[sheet_name])
        cur_row_string = str(transaction_sheet.max_row + 1)
        shipping_address = get_string_shipping_address(
            kwargs.get('shipping_address')
        )
        zipped_data = dict(
            zip(
                ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I'),  # ячейки
                [
                    kwargs.get('title'),
                    kwargs.get('price'),
                    kwargs.get('currency'),
                    kwargs.get('name'),
                    kwargs.get('email'),
                    kwargs.get('phone_number'),
                    shipping_address,
                    kwargs.get('status'),
                    kwargs.get('datetime')
                ]
            )
        )
        for i in zipped_data:
            transaction_sheet[i + cur_row_string] = zipped_data.get(i)
        workbook.save(xlsx_file_path)
        workbook.close()


def calculate_total_marge(xlsx_file_path, period=1, total=False):
    """Считает выручку за период или суммурную."""
    if total:
        if os.path.exists(path=xlsx_file_path):
            workbook = load_workbook(xlsx_file_path)
            if 'transactions' in workbook.sheetnames:
                sheet = workbook['transactions']
                max_row = sheet.max_row
                cur_amount = 0
                for price in sheet.iter_rows(
                        max_row=max_row, min_row=2, max_col=2, min_col=2,
                        values_only=True):
                    if price[0]:
                        cur_amount += price[0]
                return cur_amount
            return 0